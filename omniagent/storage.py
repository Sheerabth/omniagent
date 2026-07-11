"""MinIO storage client — session-scoped file operations.

All paths are prefixed ``sessions/{session_id}/`` automatically so callers
pass bare relative paths (``report.pdf``, ``outputs/data.csv``).
"""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING

from minio import Minio
from minio.deleteobjects import DeleteObject

from omniagent.api.models import FileRef
from omniagent.config import settings

if TYPE_CHECKING:
    pass

logger = logging.getLogger(__name__)

_SESSION_PREFIX = "sessions"


def _session_prefix(session_id: str) -> str:
    return f"{_SESSION_PREFIX}/{session_id}/"


def _object_key(session_id: str, path: str) -> str:
    prefix = _session_prefix(session_id)
    return f"{prefix}{path.lstrip('/')}"


class StorageClient:
    """MinIO client for session-scoped file operations."""

    def __init__(self) -> None:
        self._client = Minio(
            endpoint=settings.omniagent_minio_endpoint,
            access_key=settings.omniagent_minio_access_key,
            secret_key=settings.omniagent_minio_secret_key,
            secure=settings.omniagent_minio_secure,
        )
        self._bucket = settings.omniagent_minio_bucket
        self._max_read_bytes = settings.omniagent_file_read_max_bytes
        self._max_read_lines = settings.omniagent_file_read_max_lines
        self._max_upload = settings.omniagent_max_upload_size_mb * 1024 * 1024

    async def ensure_bucket(self) -> None:
        """Create bucket if it doesn't exist. Call once at startup."""
        import asyncio

        loop = asyncio.get_running_loop()
        exists = await loop.run_in_executor(None, self._client.bucket_exists, self._bucket)
        if not exists:
            await loop.run_in_executor(None, self._client.make_bucket, self._bucket)
            logger.info("created MinIO bucket %r", self._bucket)

    async def upload(
        self,
        session_id: str,
        path: str,
        data: bytes,
        content_type: str,
    ) -> FileRef:
        """Upload a file to the session. Rejects oversized uploads."""
        size = len(data)
        if size > self._max_upload:
            raise ValueError(f"file too large: {size} bytes (max {self._max_upload})")

        import asyncio

        key = _object_key(session_id, path)
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(
            None,
            lambda: self._client.put_object(
                self._bucket,
                key,
                io.BytesIO(data),
                size,
                content_type=content_type,
            ),
        )

        return FileRef(
            path=path,
            name=path.rsplit("/", 1)[-1] if "/" in path else path,
            content_type=content_type,
            size=size,
            updated_at="",  # caller stamps with current time
        )

    async def download(self, session_id: str, path: str) -> bytes:
        """Download file contents from session storage."""
        import asyncio

        key = _object_key(session_id, path)
        loop = asyncio.get_running_loop()
        response = await loop.run_in_executor(
            None, lambda: self._client.get_object(self._bucket, key)
        )
        try:
            return response.read()
        finally:
            response.close()
            response.release_conn()

    async def delete(self, session_id: str, path: str) -> None:
        """Delete a single file from session storage."""
        import asyncio

        key = _object_key(session_id, path)
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, lambda: self._client.remove_object(self._bucket, key))

    async def delete_prefix(self, session_id: str) -> None:
        """Delete all files under a session prefix. Used on session delete."""
        import asyncio

        prefix = _session_prefix(session_id)
        loop = asyncio.get_running_loop()
        objects = await loop.run_in_executor(
            None,
            lambda: list(self._client.list_objects(self._bucket, prefix=prefix, recursive=True)),
        )
        if not objects:
            return
        delete_list = [DeleteObject(obj.object_name) for obj in objects if obj.object_name]
        errors = await loop.run_in_executor(
            None,
            lambda: self._client.remove_objects(self._bucket, delete_list),
        )
        for err in errors:
            logger.warning("MinIO delete error: %s (session=%s)", err, session_id)

    async def list_objects(
        self,
        session_id: str,
        prefix: str = "",
        max_results: int = 200,
    ) -> list[FileRef]:
        """List files in session. Optional sub-prefix filter."""
        import asyncio

        search_prefix = _object_key(session_id, prefix) if prefix else _session_prefix(session_id)
        loop = asyncio.get_running_loop()
        objects = await loop.run_in_executor(
            None,
            lambda: list(
                self._client.list_objects(
                    self._bucket,
                    prefix=search_prefix,
                    recursive=True,
                )
            )[:max_results],
        )
        base = _session_prefix(session_id)
        results: list[FileRef] = []
        for obj in objects:
            if not obj.object_name:
                continue
            rel_path = obj.object_name[len(base) :]
            updated = obj.last_modified.isoformat() if obj.last_modified else ""
            results.append(
                FileRef(
                    path=rel_path,
                    name=rel_path.rsplit("/", 1)[-1] if "/" in rel_path else rel_path,
                    content_type=obj.content_type or "application/octet-stream",
                    size=obj.size or 0,
                    updated_at=updated,
                )
            )
        return results

    async def stat(self, session_id: str, path: str) -> FileRef:
        """Get file metadata without downloading content."""
        import asyncio

        key = _object_key(session_id, path)
        loop = asyncio.get_running_loop()
        obj = await loop.run_in_executor(None, lambda: self._client.stat_object(self._bucket, key))
        updated = obj.last_modified.isoformat() if obj.last_modified else ""
        return FileRef(
            path=path,
            name=path.rsplit("/", 1)[-1] if "/" in path else path,
            content_type=obj.content_type or "application/octet-stream",
            size=obj.size or 0,
            updated_at=updated,
        )

    async def file_read(
        self,
        session_id: str,
        path: str,
        offset: int = 0,
        limit: int | None = None,
        tail: int | None = None,
        grep: str | None = None,
    ) -> str:
        """Read file with surgical extraction params.

        Returns extracted text for documents, base64 for unsupported binary.
        offset/limit/tail/grep: surgical line-based extraction for text files.
        """
        data = await self.download(session_id, path)
        content_type = "application/octet-stream"

        # Determine content type from stat if we need it
        try:
            ref = await self.stat(session_id, path)
            content_type = ref.content_type
        except Exception:
            pass

        is_text = content_type.startswith("text/") or content_type in (
            "application/json",
            "application/xml",
            "application/javascript",
            "application/x-yaml",
        )

        if not is_text:
            # Server-side extraction for common document formats.
            extracted = _extract_document(data, content_type, path)
            if extracted is not None:
                return extracted

            import base64

            return base64.b64encode(data).decode("ascii")

        text = data.decode("utf-8", errors="replace")

        # Apply surgical params
        lines = text.split("\n")

        if tail is not None:
            lines = lines[-tail:]
        elif offset > 0 or limit is not None:
            start = offset
            end = offset + limit if limit is not None else len(lines)
            lines = lines[start:end]
        elif grep is not None:
            lines = [line for line in lines if grep in line]

        result = "\n".join(lines)

        # Cap output
        if len(result) > self._max_read_bytes:
            result = result[: self._max_read_bytes] + (
                f"\n\n[... truncated at {self._max_read_bytes // (1024 * 1024)}MB; use offset/limit for more ...]"
            )

        if len(lines) > self._max_read_lines:
            result = (
                "\n".join(lines[: self._max_read_lines])
                + f"\n\n[... truncated at {self._max_read_lines} lines; use offset/limit for more ...]"
            )

        return result


# ── Document extraction helpers ───────────────────────────────────────────


def _extract_document(data: bytes, content_type: str, path: str) -> str | None:
    """Extract text from common document formats. Returns None if unsupported."""
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""

    if content_type == "application/pdf" or ext == "pdf":
        return _extract_pdf(data)

    if content_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or ext in ("docx", "doc"):
        return _extract_docx(data)

    if content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or ext in ("xlsx", "xls"):
        return _extract_xlsx(data)

    return None


def _extract_pdf(data: bytes) -> str:
    """Extract text from PDF using pypdf."""
    import io

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text:
            parts.append(f"--- Page {i + 1} ---\n{text}")
    return "\n\n".join(parts) if parts else "[PDF contains no extractable text]"


def _extract_docx(data: bytes) -> str:
    """Extract text from DOCX using python-docx."""
    import io

    from docx import Document

    doc = Document(io.BytesIO(data))
    parts: list[str] = []
    for para in doc.paragraphs:
        if para.text.strip():
            parts.append(para.text)
    return "\n".join(parts) if parts else "[DOCX contains no extractable text]"


def _extract_xlsx(data: bytes) -> str:
    """Extract text from XLSX using openpyxl — renders as CSV-like text."""
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            parts.append(",".join(str(cell) if cell is not None else "" for cell in row))
    wb.close()
    return "\n".join(parts) if parts else "[XLSX contains no data]"
